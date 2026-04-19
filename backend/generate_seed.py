import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def slug(name):
    return re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')

# ── Data ──────────────────────────────────────────────────────────────────────

RECIPES = [
    # Breakfast
    {"name": "Poha",             "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/f/f4/Poha%2C_a_snack_made_of_flattened_rice.jpg",
     "url": "https://en.wikipedia.org/wiki/Poha"},
    {"name": "Upma",             "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/8/86/A_photo_of_Upma.jpg/960px-A_photo_of_Upma.jpg",
     "url": "https://en.wikipedia.org/wiki/Upma"},
    {"name": "Aloo Paratha",     "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/1/1e/Triangle_paratha_%28cropped%29.JPG/960px-Triangle_paratha_%28cropped%29.JPG",
     "url": "https://en.wikipedia.org/wiki/Aloo_paratha"},
    {"name": "Idli Sambar",      "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/1/11/Idli_Sambar.JPG",
     "url": "https://en.wikipedia.org/wiki/Idli"},
    {"name": "Dosa",             "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/4/4d/Masala_Dosa_with_Chutney%2C_Dosa.jpg",
     "url": "https://en.wikipedia.org/wiki/Dosa"},
    {"name": "Paratha + Curd",   "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/1/1e/Triangle_paratha_%28cropped%29.JPG/960px-Triangle_paratha_%28cropped%29.JPG",
     "url": "https://en.wikipedia.org/wiki/Paratha"},
    {"name": "Bread Butter Jam", "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/e/e4/Bread%2C_jam_and_butter_West_Palm_Beach.jpg",
     "url": "https://en.wikipedia.org/wiki/Toast"},
    {"name": "Besan Chilla",     "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/0/0c/Chilla_besan.JPG",
     "url": "https://en.wikipedia.org/wiki/Chilla_(food)"},
    {"name": "Chole Bhature",    "meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/9/9e/Chole_Bhature_from_Nagpur.JPG/960px-Chole_Bhature_from_Nagpur.JPG",
     "url": "https://en.wikipedia.org/wiki/Chole_bhature"},
    {"name": "Cornflakes + Milk","meal_time": "breakfast", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/c/c3/Kellogg%27s_Corn_Flakes%2C_with_milk.jpg/960px-Kellogg%27s_Corn_Flakes%2C_with_milk.jpg",
     "url": "https://en.wikipedia.org/wiki/Corn_flakes"},

    # Lunch
    {"name": "Dal Tadka",            "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/2/27/Dal_tadka_with_hot_chilli.jpg",
     "url": "https://en.wikipedia.org/wiki/Dal_tadka"},
    {"name": "Rajma",                "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/3/37/Rajma_Masala_%2832081557778%29.jpg/960px-Rajma_Masala_%2832081557778%29.jpg",
     "url": "https://en.wikipedia.org/wiki/Rajma"},
    {"name": "Chole Masala",         "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/8/8e/Chana_masala.jpg/960px-Chana_masala.jpg",
     "url": "https://en.wikipedia.org/wiki/Chana_masala"},
    {"name": "Kadhi Pakora",         "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/6/67/Kadhi_Pakora.jpg/960px-Kadhi_Pakora.jpg",
     "url": "https://en.wikipedia.org/wiki/Kadhi"},
    {"name": "Sambar",               "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/b/bb/Pumpkin_sambar.JPG/960px-Pumpkin_sambar.JPG",
     "url": "https://en.wikipedia.org/wiki/Sambar_(dish)"},
    {"name": "Dal Makhani",          "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/6/69/Punjabi_style_Dal_Makhani.jpg/960px-Punjabi_style_Dal_Makhani.jpg",
     "url": "https://en.wikipedia.org/wiki/Dal_makhani"},
    {"name": "Paneer Butter Masala", "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/5/5c/Paneer_Makhani_Veggie.jpeg",
     "url": "https://en.wikipedia.org/wiki/Paneer_butter_masala"},
    {"name": "Aloo Matar",           "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/e/ec/Aalu_Matar_Curry.jpg",
     "url": "https://en.wikipedia.org/wiki/Aloo_matar"},
    {"name": "Mix Veg",              "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/9/92/Mixed_Vegetable_Curry_-_Kolkata_2014-02-13_2644.JPG",
     "url": "https://en.wikipedia.org/wiki/Vegetable_curry"},
    {"name": "Palak Paneer",         "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/b/b7/Palakpaneer_Rayagada_Odisha_0009.jpg/960px-Palakpaneer_Rayagada_Odisha_0009.jpg",
     "url": "https://en.wikipedia.org/wiki/Palak_paneer"},
    {"name": "Jeera Rice",           "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/5/5e/Jeera-rice.JPG",
     "url": "https://en.wikipedia.org/wiki/Jeera_rice"},
    {"name": "Veg Pulao",            "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/4/41/Veg_Rice_Pulao_1.jpg",
     "url": "https://en.wikipedia.org/wiki/Pilaf"},
    {"name": "Raita",                "meal_time": "lunch", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/7/78/Cucumber-raita.jpg/960px-Cucumber-raita.jpg",
     "url": "https://en.wikipedia.org/wiki/Raita"},

    # Dinner
    {"name": "Aloo Gobi",           "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/c/c8/Aloo_gobi.jpg",
     "url": "https://en.wikipedia.org/wiki/Aloo_gobi"},
    {"name": "Baingan Bharta",      "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/4/4d/Baigan_Bharta_from_Nagpur.JPG/960px-Baigan_Bharta_from_Nagpur.JPG",
     "url": "https://en.wikipedia.org/wiki/Baingan_bharta"},
    {"name": "Dal Fry",             "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/b/b1/Dal_Fry_Recipe_In_Dhaba_Style_From_Indian_Cuisine_By_Sonia_Goyal.jpg",
     "url": "https://en.wikipedia.org/wiki/Dal_fry"},
    {"name": "Paneer Tikka Masala", "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/b/b0/Panner_tikka_Masala.JPG",
     "url": "https://en.wikipedia.org/wiki/Paneer_tikka_masala"},
    {"name": "Chana Dal Curry",     "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/e/e5/Chane_ki_daal_%289685959450%29.jpg",
     "url": "https://en.wikipedia.org/wiki/Chana_dal"},
    {"name": "Lauki Kofta",        "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/b/bf/Lauki_Kofta_Curry.JPG",
     "url": "https://en.wikipedia.org/wiki/Kofta"},
    {"name": "Bhindi Masala",       "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/2/22/Bhindi_Masala.jpg",
     "url": "https://en.wikipedia.org/wiki/Bhindi_masala"},
    {"name": "Mushroom Matar",      "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/f/f8/Matar_mushroom.jpg",
     "url": "https://en.wikipedia.org/wiki/Matar_paneer"},
    {"name": "Plain Rice",          "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/d/d6/Meshi_001.jpg/960px-Meshi_001.jpg",
     "url": "https://en.wikipedia.org/wiki/Cooked_rice"},
    {"name": "Roti",                "meal_time": "dinner", "diet_types": ["veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/f/fe/2_Chapati_warm_and_ready_to_be_eaten.jpg/960px-2_Chapati_warm_and_ready_to_be_eaten.jpg",
     "url": "https://en.wikipedia.org/wiki/Roti"},

    # Protein addons (snack meal_time)
    {"name": "Egg Bhurji",     "meal_time": "snack", "diet_types": ["non_veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/0/06/Spicy_egg_bhurji_%40_the_eggfactory.jpg/960px-Spicy_egg_bhurji_%40_the_eggfactory.jpg",
     "url": "https://en.wikipedia.org/wiki/Egg_bhurji"},
    {"name": "Boiled Eggs",    "meal_time": "snack", "diet_types": ["non_veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/c/cc/Soft-boiled-egg.jpg/960px-Soft-boiled-egg.jpg",
     "url": "https://en.wikipedia.org/wiki/Boiled_egg"},
    {"name": "Omelette",       "meal_time": "snack", "diet_types": ["non_veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/thumb/4/4d/Gorgonzola_%2B_Bacon_Omelette_%40_Omelegg_%40_Amsterdam_%2816600947041%29.jpg/960px-Gorgonzola_%2B_Bacon_Omelette_%40_Omelegg_%40_Amsterdam_%2816600947041%29.jpg",
     "url": "https://en.wikipedia.org/wiki/Omelette"},
    {"name": "Chicken Curry",  "meal_time": "snack", "diet_types": ["non_veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/f/fd/Chicken_tikka_masala.jpg",
     "url": "https://en.wikipedia.org/wiki/Chicken_curry"},
    {"name": "Keema Masala",   "meal_time": "snack", "diet_types": ["non_veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/2/29/Keema_Matar_%28a_dish_from_India%29.jpg",
     "url": "https://en.wikipedia.org/wiki/Keema"},
    {"name": "Egg Curry",      "meal_time": "snack", "diet_types": ["non_veg"],
     "photo": "https://upload.wikimedia.org/wikipedia/commons/f/fb/Egg_curry_Indian_style.jpg",
     "url": "https://en.wikipedia.org/wiki/Egg_curry"},
]

# ── Excel ─────────────────────────────────────────────────────────────────────

def styled_header(ws, headers, header_fill_hex="1F4E79"):
    fill = PatternFill("solid", fgColor=header_fill_hex)
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(style="medium"),
        right=Side(style="thin"),
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 22


def auto_width(ws, min_w=12, max_w=60):
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(length + 2, max_w))


wb = openpyxl.Workbook()

# ── Sheet: Recipes ─────────────────────────────────────────────────────────────
ws_recipes = wb.active
ws_recipes.title = "Recipes"

headers = ["ID", "Name", "Meal Time", "Diet Types", "Photo URL", "Recipe URL"]
styled_header(ws_recipes, headers)

row_fill_veg    = PatternFill("solid", fgColor="E8F5E9")
row_fill_nonveg = PatternFill("solid", fgColor="FCE4EC")

for i, r in enumerate(RECIPES, 2):
    recipe_id    = slug(r["name"])
    diet_str     = ", ".join(r["diet_types"])
    fill = row_fill_veg if "veg" in r["diet_types"] and "non_veg" not in r["diet_types"] else row_fill_nonveg

    values = [recipe_id, r["name"], r["meal_time"], diet_str, r["photo"], r["url"]]
    for col_idx, val in enumerate(values, 1):
        cell = ws_recipes.cell(row=i, column=col_idx, value=val)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=False)

ws_recipes.freeze_panes = "A2"
auto_width(ws_recipes)

# ── Sheet: SQL Inserts ─────────────────────────────────────────────────────────
ws_sql = wb.create_sheet("SQL Inserts")
ws_sql.column_dimensions["A"].width = 160

sql_lines = [
    "-- Recipe INSERT statements",
    "-- diet_types stored as PostgreSQL TEXT[] array",
    "",
]
for r in RECIPES:
    recipe_id = slug(r["name"])
    diet_arr  = "{" + ",".join(r["diet_types"]) + "}"
    photo     = r["photo"].replace("'", "''")
    url       = r["url"].replace("'", "''")
    line = (
        f"INSERT INTO recipe (id, name, meal_time, diet_types, photo, url) VALUES "
        f"('{recipe_id}', '{r['name']}', '{r['meal_time']}', '{diet_arr}', '{photo}', '{url}');"
    )
    sql_lines.append(line)

for i, line in enumerate(sql_lines, 1):
    cell = ws_sql.cell(row=i, column=1, value=line)
    if line.startswith("--"):
        cell.font = Font(color="888888", italic=True)
    elif line.startswith("INSERT"):
        cell.font = Font(name="Courier New", size=9)

wb.save("/Users/drinkprime/Desktop/bite-buddy/backend/bite_buddy_seed.xlsx")
print(f"✓ Excel saved: bite_buddy_seed.xlsx  ({len(RECIPES)} recipes)")

# ── SQL file ──────────────────────────────────────────────────────────────────
with open("/Users/drinkprime/Desktop/bite-buddy/backend/insert_recipes.sql", "w") as f:
    f.write("\n".join(sql_lines) + "\n")
print(f"✓ SQL saved:   insert_recipes.sql")
